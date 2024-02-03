import re
import urllib.parse
# from email.mime import image
import os
import sys
import json
from time import sleep
# from urllib.parse import quote
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
# import chromedriver_autoinstaller
from models.store import Store
from models.brand import Brand
from models.product import Product
from models.variant import Variant
from models.metafields import Metafields
import glob
import requests
from datetime import datetime
# import threading

from openpyxl import Workbook
from openpyxl.drawing.image import Image as Imag
# from openpyxl.utils import get_column_letter
from PIL import Image

from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# class myScrapingThread(threading.Thread):
#     def __init__(self, threadID: int, name: str, obj, username: str, brand: str, brand_code: str, product_number: str, glasses_type: str, headers: dict) -> None:
#         threading.Thread.__init__(self)
#         self.threadID = threadID
#         self.name = name
#         self.username = username
#         self.brand = brand
#         self.brand_code = brand_code
#         self.product_number = product_number
#         self.glasses_type = glasses_type
#         self.headers = headers
#         self.obj = obj
#         self.status = 'in progress'
#         pass

#     def run(self):
#         self.obj.scrape_product(self.username, self.brand, self.brand_code, self.product_number, self.glasses_type, self.headers)
#         self.status = 'completed'

#     def active_threads(self):
#         return threading.activeCount()

class Digitalhub_Scraper:
    def __init__(self, DEBUG: bool, result_filename: str, logs_filename: str) -> None:
        self.DEBUG = DEBUG
        self.data = []
        self.result_filename = result_filename
        self.logs_filename = logs_filename
        self.thread_list = []
        self.thread_counter = 0
        self.chrome_options = Options()
        self.chrome_options.add_argument('--disable-infobars')
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.browser = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=self.chrome_options)
        self.auth_token: str = ''
        pass

    def controller(self, store: Store, brands_with_types: list[dict]) -> None:
        try:
            cookies: dict = dict()
            brand_data: dict = dict()

            self.browser.get(store.link)
            self.wait_until_browsing()

            if self.login(store.username, store.password):
                if self.wait_until_element_found(20, 'xpath', '//button[text()="BRANDS"]'):
                    
                    for brand_with_type in brands_with_types:
                        brand: str = brand_with_type['brand']
                        brand_code: str = str(brand_with_type['code']).strip()
                        print(f'Brand: {brand}')

                        for glasses_type in brand_with_type['glasses_type']:

                            if not cookies: cookies = self.get_cookies()
                            if not brand_data: brand_data = self.get_brand_data(brand, cookies)

                            if brand_data:
                                brand_json = self.get_brand_url(brand, glasses_type, brand_data)

                                if brand_json:
                                    brand_url = brand_json['url']
                                    brand_code = brand_json['code']
                                    brand_category_value = brand_json['category_value']

                                    self.open_new_tab(f'{brand_url}?cleanFilters')
                                    self.wait_until_browsing()
                                    start_time = datetime.now()

                                    if self.wait_until_element_found(90, 'xpath', '//div[@c-searchlayout_searchlayout and contains(@class, "grid")]/div'):
                                        total_products = self.get_total_products()
                                        scraped_products = 0

                                        print(f'Type: {glasses_type} | Total products: {total_products}')
                                        print(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

                                        self.print_logs(f'Type: {glasses_type} | Total products: {total_products}')
                                        self.print_logs(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

                                        if total_products and int(total_products) > 0: 
                                            self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                                        page_no = 0
                                        while int(scraped_products) < int(total_products):
                                            brand_url = self.browser.current_url

                                            for product in self.get_products(brand_url, brand_code, brand_category_value, page_no, cookies):
                                                scraped_products += 1

                                                product_number = product['product_number']
                                                product_url = product['product_url']

                                                self.scrape_product(brand, product_number, product_url, glasses_type, cookies)

                                                if total_products and int(total_products) > 0: 
                                                    self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                                            page_no += 1
                                            self.save_to_json(self.data)
                                        
                                    self.save_to_json(self.data)
                                    end_time = datetime.now()

                                    print(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                    print('Duration: {}\n'.format(end_time - start_time))

                                    self.print_logs(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                    self.print_logs('Duration: {}\n'.format(end_time - start_time))

                                    self.close_last_tab()

            else: print(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')
        except Exception as e:
            if self.DEBUG: print(f'Exception in Digitalhub_Scraper controller: {e}')
            self.print_logs(f'Exception in Digitalhub_Scraper controller: {e}')
        finally: 
            self.browser.quit()
            self.wait_for_thread_list_to_complete()
            self.save_to_json(self.data)

    def wait_until_browsing(self) -> None:
        while True:
            try:
                state = self.browser.execute_script('return document.readyState; ')
                if 'complete' == state: break
                else: sleep(0.2)
            except: pass

    def login(self, username: str, password: str) -> bool:
        login_flag = False
        try:
            if self.wait_until_element_found(50, 'xpath', '//input[@id="username"]'):
                try:
                    button = WebDriverWait(self.browser, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Accept"]')))
                    button.click()
                except: pass

                self.browser.find_element(By.XPATH, '//input[@id="username"]').send_keys(username)
                self.browser.find_element(By.XPATH, '//input[@id="password"]').send_keys(password)
                try:
                    button = WebDriverWait(self.browser, 50).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Login"]')))
                    button.click()

                    WebDriverWait(self.browser, 50).until(EC.presence_of_element_located((By.XPATH, '//button[text()="BRANDS"]')))
                    login_flag = True
                except Exception as e:
                    self.print_logs(str(e))
                    if self.DEBUG: print(str(e))
                    else: pass
        except Exception as e:
            self.print_logs(f'Exception in login: {str(e)}')
            if self.DEBUG: print(f'Exception in login: {str(e)}')
            else: pass
        finally: return login_flag

    def wait_until_element_found(self, wait_value: int, type: str, value: str) -> bool:
        flag = False
        try:
            if type == 'id':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.ID, value)))
                flag = True
            elif type == 'xpath':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.XPATH, value)))
                flag = True
            elif type == 'css_selector':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CSS_SELECTOR, value)))
                flag = True
            elif type == 'class_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CLASS_NAME, value)))
                flag = True
            elif type == 'tag_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.TAG_NAME, value)))
                flag = True
        except: pass
        finally: return flag

    def get_brand_url(self, brand_name: str, glasses_type: str, brand_data: dict) -> dict:
        brand_json = ''
        try:
            type = ''
            if glasses_type == 'Eyeglasses': type = 'Optical'
            else: type = glasses_type
            for children in brand_data:
                if str(children['brandLabel']).strip().lower() == str(brand_name).strip().lower():
                    code = str(children['label']).strip().replace('Model_', '').strip()
                    for glasses_type_in_json in children['children']:
                        if str(glasses_type_in_json.get('type')).strip().lower() == str(type).strip().lower():
                            brand_url = f'https://my.marcolin.com/s{glasses_type_in_json.get("link")}'
                            category_value = str(glasses_type_in_json.get('label')).strip().split('_')[-1].strip()
                            brand_json = {
                                'url': brand_url,
                                'code': code,
                                'category_value': category_value
                            }
                            break
                if brand_json: break
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brand_url: {e}')
            self.print_logs((f'Exception in get_brand_url: {e}'))
        finally: return brand_json

    def open_new_tab(self, url: str) -> None:
        # open category in new tab
        self.browser.execute_script('window.open("'+str(url)+'","_blank");')
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])
        self.wait_until_browsing()
    
    def close_last_tab(self) -> None:
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])
 
    def get_total_products(self) -> int:
        total_products = 0
        try:
            input_string = str(self.browser.find_element(By.XPATH, '//h3[@class="title-16 uppercase"]').text).strip()
            if input_string:
                match = re.search(r'\((\d+)\)', input_string)

                if match:
                    total_products = int(match.group(1))
            # total_products = int(str(self.browser.find_element(By.XPATH, '//div[@class="row mt-4 results"]/div').text).strip().split(' ')[0])
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_total_products: {e}')
            self.print_logs(f'Exception in get_total_products: {e}')
        finally: return total_products

    def get_api_headers(self, url) -> dict:
        return {
                'authority': 'my.marcolin.com',
                'accept': '*/*',
                'accept-language': 'en-US,en;q=0.9',
                'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
                'origin': 'https://my.marcolin.com',
                'referer': url,
                'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-origin',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'x-sfdc-lds-endpoints': 'ApexActionController.execute:B2BMicrositeController.getMicrositeLogoUrl, ApexActionController.execute:B2BHeaderAndNavigationController.getHeaderMenuEnhanced, ApexActionController.execute:B2BUserUtils.getUserAccount, ApexActionController.execute:B2BUtils.getAccountCanBuyProducts, ApexActionController.execute:B2BCartController.getRawCartItems, ApexActionController.execute:B2BCountrySettingsController.getCountrySettingsByFields, ApexActionController.execute:B2BWishlistUtils.isUserLoggedOnBehalf, ApexActionController.execute:B2BCartController.getWebCartInfo, ApexActionController.execute:B2BProductDetailsController.getProductEnhanced, ApexActionController.execute:B2BCategoryUtils.getSingleCategoryId, ApexActionController.execute:B2BPermissionsController.getPermissions, ApexActionController.execute:B2BCountrySettingsController.getSalesOrgSettingsByFieldsWithAccount, ApexActionController.execute:B2BAccountSalesOrgUtils.getAccountSalesOrgEnhanced, ApexActionController.execute:B2BMicrositeController.getMicrositeFooterLinks, ApexActionController.execute:B2BSocialStreamController.getPosts, ApexActionController.execute:B2BLanguageSelectorController.getAvailableLanguagePicklist',
            }

    def get_auth_token(self, url: str, cookies: dict):
        try:
            headers = {
                'authority': 'my.marcolin.com',
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'accept-language': 'en-US,en;q=0.9',
                'cache-control': 'max-age=0',
                'referer': url,
                'sec-ch-ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'document',
                'sec-fetch-mode': 'navigate',
                'sec-fetch-site': 'same-origin',
                'sec-fetch-user': '?1',
                'upgrade-insecure-requests': '1',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
            }

            response = requests.get(url=url, cookies=cookies, headers=headers, verify=False)
            if response.status_code == 200:
                for key, value in response.cookies.get_dict().items():
                    if 'Host-ERIC_PROD' in key:
                        self.auth_token = value
                        break
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_auth_token: {e}')
            self.print_logs((f'Exception in get_auth_token: {e}'))

    def get_products(self, brand_url: str, brand_code: str, brand_category_value: str, page_no: int, cookies: dict) -> list[dict]:
        products: list = list()
        try:
            headers = self.get_api_headers(brand_url)
            headers['x-sfdc-lds-endpoints'] = 'ApexActionController.execute:B2BSearchController.productSearchDynamic'

            API_ENDPOINT = 'https://my.marcolin.com/s/sfsites/aura?r=36&aura.ApexAction.execute=1'

            categoryId = str(brand_url).split('/')[-1].strip().split('?')[0]
            pageURI = str(brand_url).replace('https://my.marcolin.com', '')
            # data = {
            #     'message': '{"actions":[{"id":"214;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BSearchController","method":"productSearchDynamic","params":{"searchQuery":"{\"categoryId\":\"'+str(categoryId)+'\",\"page\":'+str(page_no)+',\"elementsPerPage\":12,\"filters\":[{\"name\":\"Brand__c\",\"type\":\"picklist\",\"values\":[\"OR\"]},{\"name\":\"ProductType__c\",\"type\":\"picklist\",\"values\":[\"S\"]}],\"sortValue\":\"MAX(CreatedDate) DESC\"}","effectiveAccountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}}]}',
            #     'aura.context': '{"mode":"PROD","fwuid":"YWYyQV90T3g3VDhySzNWUm1kcF9WUVY4bi1LdGdMbklVbHlMdER1eVVlUGcyNDYuMTUuNS0zLjAuNA","app":"siteforce:communityApp","loaded":{"APPLICATION@markup://siteforce:communityApp":"xUUH_isHmNQqCOJ9yNTV7A","COMPONENT@markup://forceCommunity:embeddedServiceSidebar":"GfuW3QajZSdMd3TyQUVuaw","COMPONENT@markup://instrumentation:o11ySecondaryLoader":"iVoI_RYCX4m4O5loBTnQfA"},"dn":[],"globals":{},"uad":false}',
            #     'aura.pageURI': pageURI,
            #     'aura.token': self.auth_token
            # }
            data = 'message=%7B%22actions%22%3A%5B%7B%22id%22%3A%22214%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BSearchController%22%2C%22method%22%3A%22productSearchDynamic%22%2C%22params%22%3A%7B%22searchQuery%22%3A%22%7B%5C%22categoryId%5C%22%3A%5C%22'+str(categoryId)+'%5C%22%2C%5C%22page%5C%22%3A'+str(page_no)+'%2C%5C%22elementsPerPage%5C%22%3A12%2C%5C%22filters%5C%22%3A%5B%7B%5C%22name%5C%22%3A%5C%22Brand__c%5C%22%2C%5C%22type%5C%22%3A%5C%22picklist%5C%22%2C%5C%22values%5C%22%3A%5B%5C%22'+str(brand_code)+'%5C%22%5D%7D%2C%7B%5C%22name%5C%22%3A%5C%22ProductType__c%5C%22%2C%5C%22type%5C%22%3A%5C%22picklist%5C%22%2C%5C%22values%5C%22%3A%5B%5C%22'+str(brand_category_value)+'%5C%22%5D%7D%5D%2C%5C%22sortValue%5C%22%3A%5C%22MAX(CreatedDate)%20DESC%5C%22%7D%22%2C%22effectiveAccountId%22%3A%220010900001z2klzAAA%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%5D%7D&aura.context=%7B%22mode%22%3A%22PROD%22%2C%22fwuid%22%3A%22YWYyQV90T3g3VDhySzNWUm1kcF9WUVY4bi1LdGdMbklVbHlMdER1eVVlUGcyNDYuMTUuNS0zLjAuNA%22%2C%22app%22%3A%22siteforce%3AcommunityApp%22%2C%22loaded%22%3A%7B%22APPLICATION%40markup%3A%2F%2Fsiteforce%3AcommunityApp%22%3A%22xUUH_isHmNQqCOJ9yNTV7A%22%2C%22COMPONENT%40markup%3A%2F%2FforceCommunity%3AembeddedServiceSidebar%22%3A%22GfuW3QajZSdMd3TyQUVuaw%22%2C%22COMPONENT%40markup%3A%2F%2Finstrumentation%3Ao11ySecondaryLoader%22%3A%22iVoI_RYCX4m4O5loBTnQfA%22%7D%2C%22dn%22%3A%5B%5D%2C%22globals%22%3A%7B%7D%2C%22uad%22%3Afalse%7D&aura.pageURI='+urllib.parse.quote_plus(str(pageURI))+'&aura.token='+self.auth_token

            response = requests.post(url=API_ENDPOINT, cookies=cookies, headers=headers, data=data, verify=False)
            if response.status_code == 200:
                for product_json in response.json()['actions'][0]['returnValue']['returnValue']['products']:
                    product_number, external_id = '', ''
                    parentId = product_json['id']
                    text = product_json['variations'][0]['fields']['Name'] if 'Name' in product_json['variations'][0]['fields'] else product_json['variations'][0]['fields']['ExternalId__c']
                    if text:
                        product_number = str(text).strip().split('@')[0].strip()
                        external_id = str(text).strip().replace('@', '').strip().lower()
                    variantId = product_json['variations'][0]['id']

                    json_data = {
                        'product_number': product_number,
                        'product_url': f'https://my.marcolin.com/s/product/{external_id}/{variantId}?isVariation=true&parentVariation={parentId}&searchType=Model'
                    }
                    if json_data not in products:
                        products.append(json_data)
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_products: {e}')
            self.print_logs((f'Exception in get_products: {e}'))
        finally: return products

    def scrape_product(self, brand_name: str, product_number: str, product_url: str, glasses_type: str, cookies: dict) -> None:

        try:
            headers = self.get_api_headers(product_url)
            headers['x-sfdc-lds-endpoints'] = 'ApexActionController.execute:B2BCacheController.cacheFiltersPageSort, ApexActionController.execute:B2BCartController.getWebCartInfo, ApexActionController.execute:B2BProductDetailsController.getProductEnhanced, ApexActionController.execute:B2BCategoryUtils.getSingleCategoryId, ApexActionController.execute:B2BPermissionsController.getPermissions, ApexActionController.execute:B2BCountrySettingsController.getCountrySettingsByFields, ApexActionController.execute:B2BUtils.getAccountCanBuyProducts, ApexActionController.execute:B2BCartController.getRawCartItems, ApexActionController.execute:B2BCountrySettingsController.getSalesOrgSettingsByFieldsWithAccount, ApexActionController.execute:B2BAccountSalesOrgUtils.getAccountSalesOrgEnhanced'
            pageURI = str(product_url).replace('https://my.marcolin.com', '')

            product_id = str(product_url).split('?')[0].split('/')[-1].strip()
            isVariation, parentVariation = '', ''
            for value in str(product_url).split('?')[-1].split('&'):
                if 'isVariation' in value:
                    isVariation = str(value).split('=')[-1].strip()
                elif 'parentVariation' in value:
                    parentVariation = str(value).split('=')[-1].strip()

            message = "message=%7B%22actions%22%3A%5B%7B%22id%22%3A%22253%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BCacheController%22%2C%22method%22%3A%22cacheFiltersPageSort%22%2C%22params%22%3A%7B%22dataToCache%22%3A%7B%22myPages_filters%22%3A%5B%7B%22name%22%3A%22Brand__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Brand%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Adidas%20Originals%22%2C%22value%22%3A%22OR%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Adidas%20Sport%22%2C%22value%22%3A%22SP%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22BMW%22%2C%22value%22%3A%22BW%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Barton%20Perreira%22%2C%22value%22%3A%22BP%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Emilio%20Pucci%22%2C%22value%22%3A%22EP%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22GCDS%22%2C%22value%22%3A%22GD%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Gant%22%2C%22value%22%3A%22GA%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Guess%22%2C%22value%22%3A%22GU%22%2C%22checked%22%3Atrue%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Guess%20By%20Marciano%22%2C%22value%22%3A%22GM%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Harley-Davidson%22%2C%22value%22%3A%22HD%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Max%26Co%22%2C%22value%22%3A%22MO%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22MaxMara%22%2C%22value%22%3A%22MM%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Moncler%22%2C%22value%22%3A%22ML%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Skechers%22%2C%22value%22%3A%22SE%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Timberland%22%2C%22value%22%3A%22TB%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Tod's%22%2C%22value%22%3A%22TO%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Tom%20Ford%22%2C%22value%22%3A%22FT%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Tom%20Ford%20Private%20Collection%22%2C%22value%22%3A%22TP%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Web%22%2C%22value%22%3A%22WE%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Zegna%22%2C%22value%22%3A%22EZ%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22selected%22%3A1%2C%22selectedString%22%3A%22(1)%22%7D%2C%7B%22name%22%3A%22ProductType__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Product%20type%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Optical%20glasses%22%2C%22value%22%3A%22V%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Sun%20glasses%22%2C%22value%22%3A%22S%22%2C%22checked%22%3Atrue%2C%22showOption%22%3Atrue%7D%5D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22selected%22%3A1%2C%22selectedString%22%3A%22(1)%22%7D%2C%7B%22name%22%3A%22Gender__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Gender%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Female%22%2C%22value%22%3A%22F%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Male%22%2C%22value%22%3A%22M%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Girl%22%2C%22value%22%3A%22G%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Boy%22%2C%22value%22%3A%22B%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Kids%22%2C%22value%22%3A%22K%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Unisex%22%2C%22value%22%3A%22U%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22selected%22%3A0%2C%22selectedString%22%3A%22%22%7D%2C%7B%22name%22%3A%22Material__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Material%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Injected%22%2C%22value%22%3A%22030%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Metal%22%2C%22value%22%3A%22010%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Plastic%22%2C%22value%22%3A%22020%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%5D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22selected%22%3A0%2C%22selectedString%22%3A%22%22%7D%2C%7B%22name%22%3A%22FamilyColor__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Color%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Atrue%2C%22values%22%3A%5B%7B%22label%22%3A%22Black%22%2C%22value%22%3A%22Black%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorBlack%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Blue%22%2C%22value%22%3A%22Blue%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorBlue%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Brown%22%2C%22value%22%3A%22Brown%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorBrown%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Green%22%2C%22value%22%3A%22Green%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorGreen%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Grey%22%2C%22value%22%3A%22Grey%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorGrey%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Red%22%2C%22value%22%3A%22Red%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2FcolorRed%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22White%22%2C%22value%22%3A%22White%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2FcolorWhite%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Yellow%22%2C%22value%22%3A%22Yellow%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2FcolorYellow%3Foid%3D00D09000005O9TvEAK%22%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22iconClass%22%3A%22color-icon%22%2C%22hasIcons%22%3Atrue%2C%22selected%22%3A0%2C%22selectedString%22%3A%22%22%7D%2C%7B%22name%22%3A%22Shape__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Shape%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Atrue%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Butterfly%22%2C%22value%22%3A%2208%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape08%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Cat%22%2C%22value%22%3A%2202%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape02%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Geometric%22%2C%22value%22%3A%2203%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape03%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Navigator%22%2C%22value%22%3A%2210%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape10%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Oval%22%2C%22value%22%3A%2204%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape04%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Pilot%22%2C%22value%22%3A%2201%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape01%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Rectangular%22%2C%22value%22%3A%2211%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape11%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Round%22%2C%22value%22%3A%2205%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape05%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Shield%22%2C%22value%22%3A%2206%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape06%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Square%22%2C%22value%22%3A%2212%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape12%3Foid%3D00D09000005O9TvEAK%22%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22iconClass%22%3A%22shape-icon%22%2C%22hasIcons%22%3Atrue%2C%22selected%22%3A0%2C%22selectedString%22%3A%22%22%7D%2C%7B%22name%22%3A%22AMeasure__c%22%2C%22noLookup%22%3Afalse%2C%22defaultValuesRange%22%3A%7B%22min%22%3A48%2C%22max%22%3A148%7D%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Size%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Atrue%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22unitOfMeasurement%22%3A%22mm%22%2C%22values%22%3A%7B%22min%22%3A48%2C%22max%22%3A148%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22inputMax%22%3A147.9%2C%22classMax%22%3A%22max_range%22%2C%22classMin%22%3A%22min_range%22%2C%22inputMin%22%3A48.1%7D%2C%7B%22name%22%3A%22BMeasure__c%22%2C%22noLookup%22%3Afalse%2C%22defaultValuesRange%22%3A%7B%22min%22%3A28.6%2C%22max%22%3A60%7D%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22B%20Measure%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Atrue%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22unitOfMeasurement%22%3A%22mm%22%2C%22values%22%3A%7B%22min%22%3A28.6%2C%22max%22%3A60%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22inputMax%22%3A59.9%2C%22classMax%22%3A%22max_range_1%22%2C%22classMin%22%3A%22min_range_1%22%2C%22inputMin%22%3A28.7%7D%2C%7B%22name%22%3A%22LensType__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Lens%20type%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Normal%22%2C%22value%22%3A%22NOR%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Polarized%22%2C%22value%22%3A%22POL%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%5D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22selected%22%3A0%2C%22selectedString%22%3A%22%22%7D%2C%7B%22name%22%3A%22FilterProtection__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Filter%20protection%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Filter%20Protection%201%22%2C%22value%22%3A%221%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Filter%20Protection%202%22%2C%22value%22%3A%222%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Filter%20Protection%202P%22%2C%22value%22%3A%222P%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Filter%20Protection%203%22%2C%22value%22%3A%223%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Filter%20Protection%203P%22%2C%22value%22%3A%223P%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%5D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22selected%22%3A0%2C%22selectedString%22%3A%22%22%7D%2C%7B%22name%22%3A%22RXCapability__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22RX%20able%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22valuesToMatch%22%3A%5B%5D%7D%2C%7B%22name%22%3A%22Flex__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Flex%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22valuesToMatch%22%3A%5B%5D%7D%2C%7B%22name%22%3A%22BestSeller%22%2C%22isPicklist%22%3Afalse%2C%22label%22%3A%22Best%20Seller%22%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22noLookup%22%3Atrue%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%2C%7B%22name%22%3A%22Availability%22%2C%22label%22%3A%22Availability%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22noLookup%22%3Atrue%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%5D%2C%22myPages_originalFilters%22%3A%5B%7B%22name%22%3A%22Brand__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Brand%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Adidas%20Originals%22%2C%22value%22%3A%22OR%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Adidas%20Sport%22%2C%22value%22%3A%22SP%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22BMW%22%2C%22value%22%3A%22BW%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Barton%20Perreira%22%2C%22value%22%3A%22BP%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Emilio%20Pucci%22%2C%22value%22%3A%22EP%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22GCDS%22%2C%22value%22%3A%22GD%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Gant%22%2C%22value%22%3A%22GA%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Guess%22%2C%22value%22%3A%22GU%22%2C%22checked%22%3Atrue%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Guess%20By%20Marciano%22%2C%22value%22%3A%22GM%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Harley-Davidson%22%2C%22value%22%3A%22HD%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Max%26Co%22%2C%22value%22%3A%22MO%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22MaxMara%22%2C%22value%22%3A%22MM%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Moncler%22%2C%22value%22%3A%22ML%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Skechers%22%2C%22value%22%3A%22SE%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Timberland%22%2C%22value%22%3A%22TB%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Tod's%22%2C%22value%22%3A%22TO%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Tom%20Ford%22%2C%22value%22%3A%22FT%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Tom%20Ford%20Private%20Collection%22%2C%22value%22%3A%22TP%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Web%22%2C%22value%22%3A%22WE%22%2C%22checked"
            message += "%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Zegna%22%2C%22value%22%3A%22EZ%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%2C%7B%22name%22%3A%22ProductType__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Product%20type%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Goggles%22%2C%22value%22%3A%22X%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Optical%20glasses%22%2C%22value%22%3A%22V%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Sun%20glasses%22%2C%22value%22%3A%22S%22%2C%22checked%22%3Atrue%2C%22showOption%22%3Atrue%7D%5D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%2C%7B%22name%22%3A%22Gender__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Gender%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Female%22%2C%22value%22%3A%22F%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Male%22%2C%22value%22%3A%22M%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Girl%22%2C%22value%22%3A%22G%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Boy%22%2C%22value%22%3A%22B%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Kids%22%2C%22value%22%3A%22K%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Teen%20Boy%22%2C%22value%22%3A%22Y%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Teen%20Girl%22%2C%22value%22%3A%22X%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Teen%20Unisex%22%2C%22value%22%3A%22Z%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Unisex%22%2C%22value%22%3A%22U%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%2C%7B%22name%22%3A%22Material__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Material%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Aluminum%22%2C%22value%22%3A%22012%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Blue%20Filter%20Injected%22%2C%22value%22%3A%22130%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Blue%20Filter%20Magnesium%22%2C%22value%22%3A%22114%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Blue%20Filter%20Metal%22%2C%22value%22%3A%22110%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Blue%20Filter%20Plastic%22%2C%22value%22%3A%22120%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Blue%20Filter%20Titanium%22%2C%22value%22%3A%22140%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Horn%22%2C%22value%22%3A%22022%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Injected%22%2C%22value%22%3A%22030%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Metal%22%2C%22value%22%3A%22010%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Plastic%22%2C%22value%22%3A%22020%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Titanium%22%2C%22value%22%3A%22040%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%2C%7B%22name%22%3A%22FamilyColor__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Color%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Atrue%2C%22values%22%3A%5B%7B%22label%22%3A%22Black%22%2C%22value%22%3A%22Black%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorBlack%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Blue%22%2C%22value%22%3A%22Blue%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorBlue%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Brown%22%2C%22value%22%3A%22Brown%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorBrown%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Green%22%2C%22value%22%3A%22Green%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorGreen%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Grey%22%2C%22value%22%3A%22Grey%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2FcolorGrey%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Red%22%2C%22value%22%3A%22Red%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2FcolorRed%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22White%22%2C%22value%22%3A%22White%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2FcolorWhite%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Yellow%22%2C%22value%22%3A%22Yellow%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2FcolorYellow%3Foid%3D00D09000005O9TvEAK%22%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22iconClass%22%3A%22color-icon%22%2C%22hasIcons%22%3Atrue%7D%2C%7B%22name%22%3A%22Shape__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Shape%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Atrue%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Browline%22%2C%22value%22%3A%2209%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape09%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Butterfly%22%2C%22value%22%3A%2208%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape08%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Cat%22%2C%22value%22%3A%2202%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape02%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Geometric%22%2C%22value%22%3A%2203%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape03%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Navigator%22%2C%22value%22%3A%2210%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%2C%22icon%22%3A%22%2Ffile-asset%2Fshape10%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Oval%22%2C%22value%22%3A%2204%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape04%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Pilot%22%2C%22value%22%3A%2201%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape01%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Rectangular%22%2C%22value%22%3A%2211%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape11%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Round%22%2C%22value%22%3A%2205%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape05%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Shield%22%2C%22value%22%3A%2206%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape06%3Foid%3D00D09000005O9TvEAK%22%7D%2C%7B%22label%22%3A%22Square%22%2C%22value%22%3A%2212%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%2C%22icon%22%3A%22%2Ffile-asset%2Fshape12%3Foid%3D00D09000005O9TvEAK%22%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22iconClass%22%3A%22shape-icon%22%2C%22hasIcons%22%3Atrue%7D%2C%7B%22name%22%3A%22AMeasure__c%22%2C%22noLookup%22%3Afalse%2C%22defaultValuesRange%22%3A%7B%22min%22%3A0%2C%22max%22%3A176%7D%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Size%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Atrue%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22unitOfMeasurement%22%3A%22mm%22%2C%22values%22%3A%7B%22min%22%3A0%2C%22max%22%3A176%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22inputMax%22%3A175.9%2C%22classMax%22%3A%22max_range%22%2C%22classMin%22%3A%22min_range%22%2C%22inputMin%22%3A0.1%7D%2C%7B%22name%22%3A%22BMeasure__c%22%2C%22noLookup%22%3Afalse%2C%22defaultValuesRange%22%3A%7B%22min%22%3A0%2C%22max%22%3A94%7D%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22B%20Measure%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Atrue%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22unitOfMeasurement%22%3A%22mm%22%2C%22values%22%3A%7B%22min%22%3A0%2C%22max%22%3A94%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22inputMax%22%3A93.9%2C%22classMax%22%3A%22max_range_1%22%2C%22classMin%22%3A%22min_range_1%22%2C%22inputMin%22%3A0.1%7D%2C%7B%22name%22%3A%22LensType__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Lens%20type%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Crystal%22%2C%22value%22%3A%22CRI%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Crystal%22%2C%22value%22%3A%22TFL%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Demo%20-%20Clear%20Lens%22%2C%22value%22%3A%22CLR%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22High%20Contrast%22%2C%22value%22%3A%22HCO%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Kolor-up%22%2C%22value%22%3A%22KUP%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Normal%22%2C%22value%22%3A%22NOR%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Optimized%22%2C%22value%22%3A%22OTM%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Photochromic%22%2C%22value%22%3A%22FOT%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Polarized%22%2C%22value%22%3A%22POL%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%2C%7B%22name%22%3A%22FilterProtection__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Filter%20protection%22%2C%22isPicklist%22%3Atrue%2C%22isCheckbox%22%3Afalse%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%5B%7B%22label%22%3A%22Filter%20Protection%200%22%2C%22value%22%3A%220%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Filter%20Protection%200%2F2%22%2C%22value%22%3A%220%2F2%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Filter%20Protection%200%2F3%22%2C%22value%22%3A%220%2F3%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Filter%20Protection%201%22%2C%22value%22%3A%221%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Filter%20Protection%201%2F2%22%2C%22value%22%3A%221%2F2%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Atrue%7D%2C%7B%22label%22%3A%22Filter%20Protection%201%2F3%22%2C%22value%22%3A%221%2F3%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%201P%22%2C%22value%22%3A%221P%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%202%22%2C%22value%22%3A%222%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%202P%22%2C%22value%22%3A%222P%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%202R0%22%2C%22value%22%3A%222R0%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%203%22%2C%22value%22%3A%223%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%203P%22%2C%22value%22%3A%223P%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%203R0%22%2C%22value%22%3A%223R0%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%203R1%22%2C%22value%22%3A%223R1%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%20S1%2FS2%22%2C%22value%22%3A%22S1%2FS2%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%20S1%2FS3%22%2C%22value%22%3A%22S1%2FS3%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%20S1%2FS4%22%2C%22value%22%3A%22S1%2FS4%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%20S2%22%2C%22value%22%3A%22S2%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%20S2%2FS4%22%2C%22value%22%3A%22S2%2FS4%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%20S3%22%2C%22value%22%3A%22S3%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%2C%7B%22label%22%3A%22Filter%20Protection%20S3RS2%2FS4%22%2C%22value%22%3A%22S3RS2%2FS4%22%2C%22checked%22%3Afalse%2C%22showOption%22%3Afalse%7D%5D%2C%22showMoreButton%22%3Atrue%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%2C%7B%22name%22%3A%22RXCapability__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22RX%20able%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22valuesToMatch%22%3A%5B%5D%7D%2C%7B%22name%22%3A%22Flex__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Flex%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22valuesToMatch%22%3A%5B%5D%7D%2C%7B%22name%22%3A%22ClipOn__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Clip%20on%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22valuesToMatch%22%3A%5B%222%22%2C%221%22%5D%7D%2C%7B%22name%22%3A%22SustainabilityInformation__c%22%2C%22noLookup%22%3Afalse%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22label%22%3A%22Sustainability%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%2C%22valuesToMatch%22%3A%5B%22ECO%22%5D%7D%2C%7B%22name%22%3A%22BestSeller%22%2C%22isPicklist%22%3Afalse%2C%22label%22%3A%22Best%20Seller%22%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22noLookup%22%3Atrue%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%2C%7B%22name%22%3A%22Availability%22%2C%22label%22%3A%22Availability%22%2C%22isPicklist%22%3Afalse%2C%22isCheckbox%22%3Atrue%2C%22isRange%22%3Afalse%2C%22isDate%22%3Afalse%2C%22noLookup%22%3Atrue%2C%22values%22%3A%7B%22isChecked%22%3Afalse%7D%2C%22isOpen%22%3Afalse%2C%22iconName%22%3A%22utility%3Achevrondown%22%2C%22isShape%22%3Afalse%2C%22isFamilyColor%22%3Afalse%2C%22showMoreButton%22%3Afalse%2C%22showLessButton%22%3Afalse%2C%22hasIcons%22%3Afalse%7D%5D%2C%22searchResult_currentFilters%22%3A%5B%7B%22type%22%3A%22picklist%22%2C%22name%22%3A%22Brand__c%22%2C%22values%22%3A%5B%22GU%22%5D%7D%2C%7B%22type%22%3A%22picklist%22%2C%22name%22%3A%22ProductType__c%22%2C%22values"
            message += "%22%3A%5B%22S%22%5D%7D%5D%7D%2C%22page%22%3A%22PLP%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%2C%7B%22id%22%3A%22254%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BCartController%22%2C%22method%22%3A%22getWebCartInfo%22%2C%22params%22%3A%7B%22accountId%22%3A%220010900001z2klzAAA%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%2C%7B%22id%22%3A%22255%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BProductDetailsController%22%2C%22method%22%3A%22getProductEnhanced%22%2C%22params%22%3A%7B%22queryInfo%22%3A%7B%22productId%22%3A%22"+str(product_id)+"%22%2C%22isVariation%22%3A%22"+str(isVariation)+"%22%2C%22parentVariation%22%3A%22+"+str(parentVariation)+"%22%7D%2C%22accountId%22%3A%220010900001z2klzAAA%22%2C%22searchType%22%3A%22Model%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%2C%7B%22id%22%3A%22256%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BCategoryUtils%22%2C%22method%22%3A%22getSingleCategoryId%22%2C%22params%22%3A%7B%22category%22%3A%22Model%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%2C%7B%22id%22%3A%22257%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BPermissionsController%22%2C%22method%22%3A%22getPermissions%22%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%2C%7B%22id%22%3A%22258%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BCountrySettingsController%22%2C%22method%22%3A%22getCountrySettingsByFields%22%2C%22params%22%3A%7B%22accountId%22%3A%220010900001z2klzAAA%22%2C%22fieldsToQuery%22%3A%22MaxQuantityPerSKU__c%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%2C%7B%22id%22%3A%22259%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BUtils%22%2C%22method%22%3A%22getAccountCanBuyProducts%22%2C%22params%22%3A%7B%22accountId%22%3A%220010900001z2klzAAA%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%2C%7B%22id%22%3A%22260%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BCartController%22%2C%22method%22%3A%22getRawCartItems%22%2C%22params%22%3A%7B%22effectiveAccountId%22%3A%220010900001z2klzAAA%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%2C%7B%22id%22%3A%22261%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BCountrySettingsController%22%2C%22method%22%3A%22getSalesOrgSettingsByFieldsWithAccount%22%2C%22params%22%3A%7B%22accountId%22%3A%220010900001z2klzAAA%22%2C%22fieldsToQuery%22%3A%22CanShowDate__c%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%2C%7B%22id%22%3A%22262%3Ba%22%2C%22descriptor%22%3A%22aura%3A%2F%2FApexActionController%2FACTION%24execute%22%2C%22callingDescriptor%22%3A%22UNKNOWN%22%2C%22params%22%3A%7B%22namespace%22%3A%22%22%2C%22classname%22%3A%22B2BAccountSalesOrgUtils%22%2C%22method%22%3A%22getAccountSalesOrgEnhanced%22%2C%22params%22%3A%7B%22accountId%22%3A%220010900001z2klzAAA%22%7D%2C%22cacheable%22%3Afalse%2C%22isContinuation%22%3Afalse%7D%7D%5D%7D"

            data = message+"&aura.context=%7B%22mode%22%3A%22PROD%22%2C%22fwuid%22%3A%22YWYyQV90T3g3VDhySzNWUm1kcF9WUVY4bi1LdGdMbklVbHlMdER1eVVlUGcyNDYuMTUuNS0zLjAuNA%22%2C%22app%22%3A%22siteforce%3AcommunityApp%22%2C%22loaded%22%3A%7B%22APPLICATION%40markup%3A%2F%2Fsiteforce%3AcommunityApp%22%3A%22xUUH_isHmNQqCOJ9yNTV7A%22%2C%22COMPONENT%40markup%3A%2F%2FforceCommunity%3AembeddedServiceSidebar%22%3A%22GfuW3QajZSdMd3TyQUVuaw%22%2C%22COMPONENT%40markup%3A%2F%2Finstrumentation%3Ao11ySecondaryLoader%22%3A%22iVoI_RYCX4m4O5loBTnQfA%22%7D%2C%22dn%22%3A%5B%5D%2C%22globals%22%3A%7B%7D%2C%22uad%22%3Afalse%7D&aura.pageURI="+urllib.parse.quote_plus(str(pageURI))+"&aura.token="+self.auth_token

            API_ENDPOINT = 'https://my.marcolin.com/s/sfsites/aura?r=54&aura.ApexAction.execute=10'

            frame_codes = []

            response = requests.post(url=API_ENDPOINT, cookies=cookies, headers=headers, data=data, verify=False)
            if response.status_code == 200:
                for value in response.json()['actions']:
                    if value['id'] == '255;a':
                        for variant_json in value['returnValue']['returnValue']:
                            frame_code = ''

                            try: frame_code = variant_json.get('fields').get('FullColorCode__c')
                            except:
                                try: frame_code = variant_json.get('fields').get('Color__c')
                                except:
                                    try: frame_code = str(variant_json.get('fields').get('ColorLabel')).split('-')[0].strip()
                                    except: pass

                            if frame_code not in frame_codes:
                                frame_codes.append(frame_code)
                                try:
                                    product = Product()

                                    product.brand = str(brand_name).strip()
                                    try: product.number = variant_json.get('fields').get('ModelCode__c')
                                    except: pass
                                    try: 
                                        name = variant_json.get('fields').get('Name') 
                                        product.name = name if name and '@' not in name else ''
                                    except: pass
                                    product.type = glasses_type
                                    product.frame_code = frame_code
                                    product.url = product_url

                                    try: product.frame_color = variant_json.get('fields').get('FrontalColor__c')
                                    except:
                                        try: product.frame_color = variant_json.get('fields').get('TempleColor__c')
                                        except: pass

                                    try: product.lens_color = variant_json.get('fields').get('LensColor__c')
                                    except: pass

                                    metafields = Metafields()

                                    try: metafields.img_url = variant_json.get('images').get('PDPImage')
                                    except: pass

                                    try: metafields.img_360_urls = variant_json.get('images').get('Rotations')
                                    except: pass

                                    try:
                                        metafields.for_who = variant_json.get('fields').get('GenderLabel')
                                        if metafields.for_who == 'Male': metafields.for_who = 'Men'
                                        elif metafields.for_who == 'Female': metafields.for_who = 'Women'
                                    except: pass

                                    try: metafields.lens_technology = variant_json.get('fields').get('LensType__c')
                                    except: pass

                                    

                                    try: metafields.frame_shape = variant_json.get('fields').get('Shape__c')
                                    except: pass

                                    try: metafields.frame_material = variant_json.get('fields').get('Material__c')
                                    except: pass


                                    try: metafields.gtin1 = variant_json.get('fields').get('EAN__c')
                                    except: pass

                                    try:
                                        size = variant_json.get('fields').get('Size__c')
                                        bridge = variant_json.get('fields').get('NoseMeasure__c')
                                        template = variant_json.get('fields').get('TempleLength__c')
                                        metafields.product_size = f'{size}-{bridge}-{template}'
                                    except: pass

                                    product.metafields = metafields

                                    variant = Variant()

                                    try: variant.title = variant_json.get('fields').get('Size__c')
                                    except: pass

                                    try: variant.sku = f'{product.number} {product.frame_code} {variant.title}'
                                    except: pass

                                    try:
                                        if str(variant_json.get('stock').get('stockColor')).strip().lower() == 'green' and str(variant_json.get('stock').get('stockLabel')).strip().title() == 'Available':
                                            variant.inventory_quantity = 5
                                        else: 
                                            variant.inventory_quantity = 0
                                    except: pass

                                    try: variant.wholesale_price = variant_json.get('prices').get('negotiatedPrice')
                                    except: pass

                                    try: variant.listing_price = variant_json.get('prices').get('retailPrice')
                                    except: pass

                                    try: variant.barcode_or_gtin = variant_json.get('fields').get('EAN__c')
                                    except: pass

                                    try: variant.size = f"{variant.title}-{variant_json.get('fields').get('NoseMeasure__c')}-{variant_json.get('fields').get('TempleLength__c')}"
                                    except: pass

                                    product.variants = variant

                                    self.data.append(product)
                                except: pass
                            else:
                                for product in self.data:
                                    if product.frame_code == frame_code and product.number == product_number:
                                        variant = Variant()

                                        try: variant.title = variant_json.get('fields').get('Size__c')
                                        except: pass

                                        try: variant.sku = f'{product.number} {product.frame_code} {variant.title}'
                                        except: pass

                                        try:
                                            if str(variant_json.get('stock').get('stockColor')).strip().lower() == 'green' and str(variant_json.get('stock').get('stockLabel')).strip().title() == 'Available':
                                                variant.inventory_quantity = 5
                                            else: variant.inventory_quantity = 0
                                        except: pass

                                        try: variant.wholesale_price = variant_json.get('prices').get('negotiatedPrice')
                                        except: pass

                                        try: variant.listing_price = variant_json.get('prices').get('retailPrice')
                                        except: pass

                                        try: variant.barcode_or_gtin = variant_json.get('fields').get('EAN__c')
                                        except: pass

                                        try: variant.size = f'{variant.title}-{product.bridge}-{product.template}'
                                        except: pass

                                        product.variants = variant

                                        try: product.metafields.gtin1 += f', {variant.barcode_or_gtin}'
                                        except: pass

                                        try:
                                            product.metafields.size_bridge_template += f', {variant.title}-{product.bridge}-{product.template}'
                                        except: pass

        except Exception as e:
            if self.DEBUG: print(f'Exception in scrape_product_data: {e}')
            self.print_logs(f'Exception in scrape_product_data: {e}')
        
    def get_cookies(self) -> dict:
        cookies: dict = {}
        try:
            for browser_cookie in self.browser.get_cookies():
                cookies[browser_cookie['name']] = browser_cookie['value']
                # # if browser_cookie["name"] == 'php-console-server':
                # #     cookies = f'{browser_cookie["name"]}={browser_cookie["value"]}; _gat_UA-153573784-1=1; {cookies}'
                # # else:
                # cookies = f'{browser_cookie["name"]}={browser_cookie["value"]}; {cookies}'
            # cookies = cookies.strip()[:-1]
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_cookies: {e}')
            self.print_logs(f'Exception in get_cookies: {e}')
        finally: return cookies

    def get_brand_data(self, brand_name: str, cookies: dict) -> dict:
        brand_data: dict = dict()
        try:
            headers = self.get_api_headers('https://my.marcolin.com/s/')
            headers['x-sfdc-lds-endpoints'] = 'ApexActionController.execute:B2BMicrositeController.getMicrositeLogoUrl, ApexActionController.execute:B2BHeaderAndNavigationController.getHeaderMenuEnhanced, ApexActionController.execute:B2BUserUtils.getUserAccount, ApexActionController.execute:B2BUtils.getAccountCanBuyProducts, ApexActionController.execute:B2BCartController.getRawCartItems, ApexActionController.execute:B2BCountrySettingsController.getCountrySettingsByFields, ApexActionController.execute:B2BWishlistUtils.isUserLoggedOnBehalf, ApexActionController.execute:B2BUtils.getSaleforceURL, ApexActionController.execute:B2BPermissionsController.getPermissions, ApexActionController.execute:B2BAccountSalesOrgUtils.getAccountSalesOrgEnhanced, ApexActionController.execute:B2BBestSellerController.getAllBestsellersEnhanced, ApexActionController.execute:B2BRecentlyViewedController.getRecentlyViewedProducts, ApexActionController.execute:B2BCountrySettingsController.getSalesOrgSettingsByFieldsWithAccount, ApexActionController.execute:B2BAnalyticsController.getUserInfoForAnalytics, ApexActionController.execute:B2BMicrositeController.getMicrositeFooterLinks, ApexActionController.execute:B2BSocialStreamController.getPosts, ApexActionController.execute:B2BLanguageSelectorController.getAvailableLanguagePicklist'


            BRAND_DATA_API = "https://my.marcolin.com/s/sfsites/aura?r=2&aura.ApexAction.execute=22&ui-communities-components-aura-components-forceCommunity-embeddedService.EmbeddedService.getStaticButtonConfigurationValues=1"
            if not self.auth_token: self.get_auth_token('https://my.marcolin.com/s/', cookies)
            data = {
                'message': '{"actions":[{"id":"88;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BMicrositeController","method":"getMicrositeLogoUrl","params":{"accountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"89;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BHeaderAndNavigationController","method":"getHeaderMenuEnhanced","params":{"accountId":"0010900001z2klzAAA","language":"en-US"},"cacheable":false,"isContinuation":false}},{"id":"90;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BUserUtils","method":"getUserAccount","params":{"effectiveAccountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"91;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BUtils","method":"getAccountCanBuyProducts","params":{"accountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"92;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BCartController","method":"getRawCartItems","params":{"effectiveAccountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"93;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BCountrySettingsController","method":"getCountrySettingsByFields","params":{"accountId":"0010900001z2klzAAA","fieldsToQuery":"MaxQuantityPerSKU__c"},"cacheable":false,"isContinuation":false}},{"id":"94;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BWishlistUtils","method":"isUserLoggedOnBehalf","cacheable":false,"isContinuation":false}},{"id":"95;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BUtils","method":"getSaleforceURL","cacheable":false,"isContinuation":false}},{"id":"96;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BPermissionsController","method":"getPermissions","cacheable":false,"isContinuation":false}},{"id":"97;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BUtils","method":"getAccountCanBuyProducts","params":{"accountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"98;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BCountrySettingsController","method":"getCountrySettingsByFields","params":{"accountId":"0010900001z2klzAAA","fieldsToQuery":"MaxQuantityPerSKU__c, CountryCode__c"},"cacheable":false,"isContinuation":false}},{"id":"99;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BAccountSalesOrgUtils","method":"getAccountSalesOrgEnhanced","params":{"accountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"100;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BBestSellerController","method":"getAllBestsellersEnhanced","params":{"accountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"101;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BRecentlyViewedController","method":"getRecentlyViewedProducts","params":{"effAccId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"102;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BCartController","method":"getRawCartItems","params":{"effectiveAccountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"103;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BCountrySettingsController","method":"getSalesOrgSettingsByFieldsWithAccount","params":{"accountId":"0010900001z2klzAAA","fieldsToQuery":"CanShowDate__c"},"cacheable":false,"isContinuation":false}},{"id":"104;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BAnalyticsController","method":"getUserInfoForAnalytics","cacheable":false,"isContinuation":false}},{"id":"109;a","descriptor":"serviceComponent://ui.communities.components.aura.components.forceCommunity.embeddedService.EmbeddedServiceController/ACTION$getStaticButtonConfigurationValues","callingDescriptor":"markup://forceCommunity:embeddedServiceFeature","params":{"eswConfigDevName":"MarcolinBotB2BPortal"},"version":"59.0"},{"id":"110;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BMicrositeController","method":"getMicrositeLogoUrl","params":{"accountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"111;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BCountrySettingsController","method":"getCountrySettingsByFields","params":{"accountId":"0010900001z2klzAAA","fieldsToQuery":"WeChat__c, WeiBo__c"},"cacheable":false,"isContinuation":false}},{"id":"112;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BMicrositeController","method":"getMicrositeFooterLinks","params":{"accountId":"0010900001z2klzAAA"},"cacheable":false,"isContinuation":false}},{"id":"113;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BSocialStreamController","method":"getPosts","cacheable":false,"isContinuation":false}},{"id":"114;a","descriptor":"aura://ApexActionController/ACTION$execute","callingDescriptor":"UNKNOWN","params":{"namespace":"","classname":"B2BLanguageSelectorController","method":"getAvailableLanguagePicklist","cacheable":false,"isContinuation":false}}]}',
                'aura.context': '{"mode":"PROD","fwuid":"YWYyQV90T3g3VDhySzNWUm1kcF9WUVY4bi1LdGdMbklVbHlMdER1eVVlUGcyNDYuMTUuNS0zLjAuNA","app":"siteforce:communityApp","loaded":{"APPLICATION@markup://siteforce:communityApp":"xUUH_isHmNQqCOJ9yNTV7A","COMPONENT@markup://forceCommunity:embeddedServiceSidebar":"GfuW3QajZSdMd3TyQUVuaw","COMPONENT@markup://instrumentation:o11ySecondaryLoader":"iVoI_RYCX4m4O5loBTnQfA"},"dn":[],"globals":{},"uad":false}',
                'aura.pageURI': '/s/',
                'aura.token': self.auth_token,
            }

            response = requests.post(url=BRAND_DATA_API, headers=headers, cookies=cookies, data=data, verify=False)
            if response.status_code == 200:
                for value in response.json()['actions']:
                    if str(brand_name).strip().title() in str(value).strip():
                        for value2 in json.loads(value['returnValue']['returnValue']):
                            if value2['section'] == 'nav':
                                brand_data = value2['children']
                                break
                    if brand_data: break
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brand_data: {e}')
            self.print_logs((f'Exception in get_brand_data: {e}'))
        finally: return brand_data
    
    def save_to_json(self, products: list[Product]) -> None:
        try:
            json_products = []
            for product in products:
                json_varinats = []
                for index, variant in enumerate(product.variants):
                    json_varinat = {
                        'position': (index + 1), 
                        'title': variant.title, 
                        'sku': variant.sku, 
                        'inventory_quantity': variant.inventory_quantity,
                        'found_status': variant.found_status,
                        'listing_price': variant.listing_price, 
                        'wholesale_price': variant.wholesale_price,
                        'barcode_or_gtin': variant.barcode_or_gtin,
                        'size': variant.size,
                        'weight': variant.weight
                    }
                    json_varinats.append(json_varinat)
                json_product = {
                    'brand': product.brand, 
                    'number': product.number, 
                    'name': product.name, 
                    'frame_code': product.frame_code, 
                    'frame_color': product.frame_color, 
                    'lens_code': product.lens_code, 
                    'lens_color': product.lens_color, 
                    'status': product.status, 
                    'type': product.type, 
                    'url': product.url, 
                    'metafields': [
                        { 'key': 'for_who', 'value': product.metafields.for_who },
                        { 'key': 'product_size', 'value': product.metafields.product_size }, 
                        { 'key': 'lens_material', 'value': product.metafields.lens_material }, 
                        { 'key': 'lens_technology', 'value': product.metafields.lens_technology }, 
                        { 'key': 'frame_material', 'value': product.metafields.frame_material }, 
                        { 'key': 'frame_shape', 'value': product.metafields.frame_shape },
                        { 'key': 'gtin1', 'value': product.metafields.gtin1 }, 
                        { 'key': 'img_url', 'value': product.metafields.img_url },
                        { 'key': 'img_360_urls', 'value': product.metafields.img_360_urls }
                    ],
                    'variants': json_varinats
                }
                json_products.append(json_product)
            
           
            with open(self.result_filename, 'w') as f: json.dump(json_products, f)
            
        except Exception as e:
            if self.DEBUG: print(f'Exception in save_to_json: {e}')
            self.print_logs(f'Exception in save_to_json: {e}')
    
    # print logs to the log file
    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass

    def printProgressBar(self, iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r") -> None:
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
            printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
        """
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
        # Print New Line on Complete
        if iteration == total: 
            print()

    # def create_thread(self, username: str, brand: str, brand_code: str, product_number: str, glasses_type: str, headers: dict) -> None:
    #     thread_name = "Thread-"+str(self.thread_counter)
    #     self.thread_list.append(myScrapingThread(self.thread_counter, thread_name, self, username, brand, brand_code, product_number, glasses_type, headers))
    #     self.thread_list[self.thread_counter].start()
    #     self.thread_counter += 1

    # def is_thread_list_complted(self) -> bool:
    #     for obj in self.thread_list:
    #         if obj.status == "in progress":
    #             return False
    #     return True

    # def wait_for_thread_list_to_complete(self) -> None:
    #     while True:
    #         result = self.is_thread_list_complted()
    #         if result: 
    #             self.thread_counter = 0
    #             self.thread_list.clear()
    #             break
    #         else: sleep(1)


def read_data_from_json_file(DEBUG, result_filename: str):
    data = []
    try:
        files = glob.glob(result_filename)
        if files:
            f = open(files[-1])
            json_data = json.loads(f.read())
            products = []

            for json_d in json_data:
                number, frame_code, brand, img_url, frame_color, lens_color = '', '', '', '', '', ''
                # product = Product()
                brand = json_d['brand']
                number = str(json_d['number']).strip().upper()
                if '/' in number: number = number.replace('/', '-').strip()
                # product.name = str(json_d['name']).strip().upper()
                frame_code = str(json_d['frame_code']).strip().upper()
                if '/' in frame_code: frame_code = frame_code.replace('/', '-').strip()
                frame_color = str(json_d['frame_color']).strip().title()
                # lens_code = str(json_d['lens_code']).strip().upper()
                lens_color = str(json_d['lens_color']).strip().title()
                # product.status = str(json_d['status']).strip().lower()
                # product.type = str(json_d['type']).strip().title()
                # product.url = str(json_d['url']).strip()
                # metafields = Metafields()
                
                for json_metafiels in json_d['metafields']:
                    # if json_metafiels['key'] == 'for_who':metafields.for_who = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'product_size':metafields.product_size = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'activity':metafields.activity = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'lens_material':metafields.lens_material = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'graduabile':metafields.graduabile = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'interest':metafields.interest = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'lens_technology':metafields.lens_technology = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'frame_material':metafields.frame_material = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'frame_shape':metafields.frame_shape = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'gtin1':metafields.gtin1 = str(json_metafiels['value']).strip().title()
                    if json_metafiels['key'] == 'img_url':img_url = str(json_metafiels['value']).strip()
                    # elif json_metafiels['key'] == 'img_360_urls':
                    #     value = str(json_metafiels['value']).strip()
                    #     if '[' in value: value = str(value).replace('[', '').strip()
                    #     if ']' in value: value = str(value).replace(']', '').strip()
                    #     if "'" in value: value = str(value).replace("'", '').strip()
                    #     for v in value.split(','):
                    #         metafields.img_360_urls = str(v).strip()
                # product.metafields = metafields
                for json_variant in json_d['variants']:
                    sku, price = '', ''
                    # variant = Variant()
                    # variant.position = json_variant['position']
                    # variant.title = str(json_variant['title']).strip()
                    sku = str(json_variant['sku']).strip().upper()
                    if '/' in sku: sku = sku.replace('/', '-').strip()
                    # variant.inventory_quantity = json_variant['inventory_quantity']
                    # variant.found_status = json_variant['found_status']
                    wholesale_price = str(json_variant['wholesale_price']).strip()
                    listing_price = str(json_variant['listing_price']).strip()
                    # variant.barcode_or_gtin = str(json_variant['barcode_or_gtin']).strip()
                    # variant.size = str(json_variant['size']).strip()
                    # variant.weight = str(json_variant['weight']).strip()
                    # product.variants = variant

                    image_attachment = download_image(img_url)
                    if image_attachment:
                        with open(f'Images/{sku}.jpg', 'wb') as f: f.write(image_attachment)
                        crop_downloaded_image(f'Images/{sku}.jpg')
                    data.append([number, frame_code, frame_color, lens_color, brand, sku, wholesale_price, listing_price])
    except Exception as e:
        if DEBUG: print(f'Exception in read_data_from_json_file: {e}')
        else: pass
    finally: return data

def download_image(url):
    image_attachment = ''
    try:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-Encoding': 'gzip, deflate, br',
            'accept-Language': 'en-US,en;q=0.9',
            'cache-Control': 'max-age=0',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'Sec-Fetch-User': '?1',
            'upgrade-insecure-requests': '1',
        }
        counter = 0
        while True:
            try:
                response = requests.get(url=url, headers=headers, timeout=20)
                # print(response.status_code)
                if response.status_code == 200:
                    # image_attachment = base64.b64encode(response.content)
                    image_attachment = response.content
                    break
                else: print(f'{response.status_code} found for downloading image')
            except: sleep(0.3)
            counter += 1
            if counter == 10: break
    except Exception as e: print(f'Exception in download_image: {str(e)}')
    finally: return image_attachment

def crop_downloaded_image(filename):
    try:
        im = Image.open(filename)
        width, height = im.size   # Get dimensions
        new_width = 1680
        new_height = 1020
        if width > new_width and height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
    except Exception as e: print(f'Exception in crop_downloaded_image: {e}')

def saving_picture_in_excel(data: list):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value='Model Code')
    worksheet.cell(row=1, column=2, value='Lens Code')
    worksheet.cell(row=1, column=3, value='Color Frame')
    worksheet.cell(row=1, column=4, value='Color Lens')
    worksheet.cell(row=1, column=5, value='Brand')
    worksheet.cell(row=1, column=6, value='SKU')
    worksheet.cell(row=1, column=7, value='Wholesale Price')
    worksheet.cell(row=1, column=8, value='Listing Price')
    worksheet.cell(row=1, column=9, value="Image")

    for index, d in enumerate(data):
        new_index = index + 2

        worksheet.cell(row=new_index, column=1, value=d[0])
        worksheet.cell(row=new_index, column=2, value=d[1])
        worksheet.cell(row=new_index, column=3, value=d[2])
        worksheet.cell(row=new_index, column=4, value=d[3])
        worksheet.cell(row=new_index, column=5, value=d[4])
        worksheet.cell(row=new_index, column=6, value=d[5])
        worksheet.cell(row=new_index, column=7, value=d[6])
        worksheet.cell(row=new_index, column=8, value=d[7])

        image = f'Images/{d[-3]}.jpg'
        if os.path.exists(image):
            im = Image.open(image)
            width, height = im.size
            worksheet.row_dimensions[new_index].height = height
            worksheet.add_image(Imag(image), anchor='I'+str(new_index))
            # col_letter = get_column_letter(9)
            # worksheet.column_dimensions[col_letter].width = width
        # print(index, image)

    workbook.save('Digitalhub Results.xlsx')

DEBUG = True
try:
    pathofpyfolder = os.path.realpath(sys.argv[0])
    # get path of Exe folder
    path = pathofpyfolder.replace(pathofpyfolder.split('\\')[-1], '')
    # download chromedriver.exe with same version and get its path
    # if os.path.exists('chromedriver.exe'): os.remove('chromedriver.exe')
    if os.path.exists('Digitalhub Results.xlsx'): os.remove('Digitalhub Results.xlsx')

    # chromedriver_autoinstaller.install(path)
    if '.exe' in pathofpyfolder.split('\\')[-1]: DEBUG = False
    
    f = open('Digitalhub start.json')
    json_data = json.loads(f.read())
    f.close()

    brands = json_data['brands']

    
    f = open('requirements/digitalhub.json')
    data = json.loads(f.read())
    f.close()

    store = Store()
    store.link = data['url']
    store.username = data['username']
    store.password = data['password']
    store.login_flag = True

    result_filename = 'requirements/Digitalhub Results.json'

    if not os.path.exists('Logs'): os.makedirs('Logs')

    log_files = glob.glob('Logs/*.txt')
    if len(log_files) > 5:
        oldest_file = min(log_files, key=os.path.getctime)
        os.remove(oldest_file)
        log_files = glob.glob('Logs/*.txt')

    scrape_time = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
    logs_filename = f'Logs/Logs {scrape_time}.txt'
    
    Digitalhub_Scraper(DEBUG, result_filename, logs_filename).controller(store, brands)
    
    for filename in glob.glob('Images/*'): os.remove(filename)
    data = read_data_from_json_file(DEBUG, result_filename)
    os.remove(result_filename)

    saving_picture_in_excel(data)
except Exception as e:
    if DEBUG: print('Exception: '+str(e))
    else: pass
